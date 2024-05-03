import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
import os
import numpy as np
import re


def number_to_string(num):
    return str(num).zfill(2)

def list_files_in_directory(directory):
    """Lists files."""
    try:
        # Get the list of files in the directory
        files = [file for file in os.listdir(directory) if os.path.isfile(os.path.join(directory, file))]

        # Sort the files alphabetically
        files_sorted = sorted(files)
        
    except FileNotFoundError:
        print(f"The directory {directory} does not exist.")
    except PermissionError:
        print(f"Permission denied to access the directory {directory}.")
    except Exception as e:
        print(f"An error occurred: {e}")
    return files_sorted

def read_eflux_file(file:str):
    # Read the file into a DataFrame, starting from line 2 and ending at line 7681, skipping the first line
    # Since we are skipping the first line (index 0), we start from index 1 (second line) and read 7680 lines
    df_eflux = pd.read_csv(file, header=None, skiprows=1, nrows=7680, delim_whitespace=True, on_bad_lines='skip')

    # Define new headers
    df_eflux.columns = ['MLT', 'ML', '[mW m^-2]']

    return df_eflux

def read_IMF_file():
    #read IMF data
    file_path_IMF = r'C:\Users\might\OneDrive - Butler Community College\Butler\Research\NASA project\SSUSI\Data\data_prep\20150317_IMF.txt'

    # Read the file into a DataFrame without headers
    df_IMF = pd.read_csv(file_path_IMF, header=None, delim_whitespace=True)

    # Define new headers
    headers = ["Year", "Month", "Day", "Hour", "Min", "Sec", "Msec", "Bx[nT]", "By[nT]", "Bz[nT]", "Vx[km/s]", "Vy[km/s]", "Vz[km/s]", "N[cm^(-3)]", "T[Kelvin]"]

    # Set the column headers
    df_IMF.columns = headers

    return df_IMF

def df_to_openpyxl(dataframe,workbook):
    # Convert the dataframe to an openpyxl object
    rows = dataframe_to_rows(dataframe, index=False, header=True)

    # Create a new workbook and worksheet
    worksheet = workbook.active
    
    # Add the rows to the worksheet
    for rnum, row in enumerate(rows, start=0):
        for col_num , the_cell in enumerate(row, start=0):
            worksheet.cell(row=rnum+1, column=col_num+1, value=the_cell)
    
    return worksheet

def extract_time(s):
    # Use a regular expression to find a pattern where digits precede "UT"
    match = re.search(r'(\d+)UT', s)
    if match:
        time_code = match.group(1)
        
        # Slice to get hour and minute
        hour = int(time_code[:2])
        minute = int(time_code[2:])
        
        return hour, minute

    # Return None if no match is found (optional, depends on how you want to handle errors)
    return None

year = '2015'
month = '03'
day = '17'
hour = 0
minute = 0

#start to read eFlux files
file_dir_eFlux = r"C:\Users\might\OneDrive - Butler Community College\Butler\Research\NASA project\SSUSI\Data\data_prep\Eflux"
file_dir_main = r"C:\Users\might\OneDrive - Butler Community College\Butler\Research\NASA project\SSUSI\Data\data_prep"

column_headers = [
    "Year", "Month", "Day", "Hour", "Min", "Sec", "Msec",
    "Bx[nT]", "By[nT]", "Bz[nT]", "Vx[km/s]", "Vy[km/s]", "Vz[km/s]",
    "N[cm^(-3)]", "T[Kelvin]", "MLT", "ML", "[mW m^-2]"
]

# Create an empty excel with these columns
wb_combined = openpyxl.Workbook()
ws_combined = wb_combined.active

# Add the headers to the first row of the worksheet
for col, header in enumerate(column_headers, start=1):
    ws_combined.cell(row=1, column=col, value=header)

#wb_IMF = openpyxl.Workbook()

df_IMF = read_IMF_file()
#ws_IMF = df_to_openpyxl(df_IMF, wb_IMF)

wb_eflux = openpyxl.Workbook()

eflux_files = list_files_in_directory(file_dir_eFlux)
for file in eflux_files:
    #print(file)
    file_path_eFlux = f"{file_dir_eFlux}\{file}" 
    hour, minute = extract_time(file)
    print("hour = ", hour, "minute = ", minute)
    imf_filtered = df_IMF[(df_IMF['Hour'] == hour) & (df_IMF['Min'] == minute) & (df_IMF['Day'] == int(day))]
    data_IFM = imf_filtered.values[0]
    #print(data_IFM)
    #print(data_IFM)
    df_eflux = read_eflux_file(file_path_eFlux)
    ws_eflux = df_to_openpyxl(df_eflux, wb_eflux)
    for row in ws_eflux.iter_rows(min_row=2):
        new_row = []
        new_row = data_IFM
        for cell in row:
            #print("cell = ",cell.value)
            new_row = np.append(new_row, cell.value)
            #print("new row= " , new_row)
        ws_combined.append(new_row.tolist())    
        

# Save the workbook to a file
wb_combined.save(fr"{file_dir_main}\test.xlsx")
wb_eflux.save(fr"{file_dir_main}\eflux_test.xlsx")